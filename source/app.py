"""
Concrestats — backend Flask (reconstruído).

Este arquivo foi RECONSTRUÍDO a partir do contrato observado do build original
(endpoints chamados pelo frontend + formato exato das respostas de /api/upload e
/api/export_report_custom verificados ao vivo). Comportamentos não observáveis
diretamente (save_file, import_merge, persistência de receitas) foram
implementados de forma sensata e estão marcados com  # [reconstruído] — confira
contra o comportamento esperado ao testar.

Compatível com PyInstaller (onedir): templates/ e static/ são localizados via
sys._MEIPASS; uploads/ e exports/ ficam ao lado do executável.

Requisitos: ver requirements.txt  (Flask 3.0.3, pandas, openpyxl, numpy).
Build:      ver Concrestats.spec / build.bat
"""

import os
import sys
import io
import json
import datetime
import re
import shutil
import threading
import time
import webbrowser

import atualizador
import licenca
import pagamento
import numpy as np
import pandas as pd
from flask import (
    Flask, request, jsonify, send_file, Response
)

# ──────────────────────────────────────────────────────────────────────────
# Localização de recursos (compatível com PyInstaller)
# ──────────────────────────────────────────────────────────────────────────
def resource_path(rel):
    """templates/ e static/ — embutidos no bundle (sys._MEIPASS quando frozen)."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel)


def app_dir():
    """Pasta GRAVÁVEL ao lado do executável (uploads/exports/receitas.json)."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


UPLOADS_DIR = os.path.join(app_dir(), "uploads")
EXPORTS_DIR = os.path.join(app_dir(), "exports")
RECEITAS_FILE = os.path.join(app_dir(), "receitas.json")
PREFS_FILE = os.path.join(app_dir(), "prefs.json")
os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(EXPORTS_DIR, exist_ok=True)

app = Flask(
    __name__,
    template_folder=resource_path("templates"),
    static_folder=resource_path("static"),
)
# Limite de upload (evita travar o app com um arquivo gigante). Planilhas
# reais têm poucos MB; 200 MB é folga de sobra.
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024

# ──────────────────────────────────────────────────────────────────────────
# Estado de sessão (em memória, por session_id)
#   SESSIONS[sid] = {
#       "sheets": { nome: {"headers": [...], "data": [[...]]} },
#       "active": nome,
#       "path":   caminho do arquivo original (para "Salvar"),
#   }
# ──────────────────────────────────────────────────────────────────────────
SESSIONS = {}

# ──────────────────────────────────────────────────────────────────────────
# MODO DE EXECUÇÃO
#   True  = servido na web (gunicorn/host) → recursos que tocam o disco do
#           servidor ficam BLOQUEADOS (ler caminho arbitrário, salvar no
#           arquivo de origem, recarregar). Sem isso, um usuário remoto
#           conseguiria ler/escrever arquivos da máquina que hospeda.
#   False = app de mesa (definido em __main__), tudo liberado.
# ──────────────────────────────────────────────────────────────────────────
MODO_WEB = os.environ.get("CONCRESTATS_WEB", "1") != "0"


def _bloqueado_na_web():
    return jsonify({
        "success": False,
        "error": ("recurso disponível apenas no aplicativo instalado — "
                  "na versão web use Abrir/Adicionar para enviar o arquivo "
                  "e Exportar para baixar o resultado")
    }), 403


def edicao():
    """Le' edicao.json, ao lado do executavel. Diz o que esta cópia mostra.

    Serve para entregar o mesmo programa com abas diferentes por cliente. E'
    decisao de EXECUCAO, nao de arquivo: se eu tirasse a aba editando o
    index.html, a primeira atualizacao automatica traria ela de volta, porque
    o pacote substitui templates/ e static/ inteiros. Lido aqui, sobrevive a
    qualquer atualizacao.

    Sem o arquivo, nada e' escondido — que e' o comportamento de sempre.
    """
    try:
        with open(os.path.join(app_dir(), "edicao.json"), encoding="utf-8") as fh:
            d = json.load(fh) or {}
    except Exception:  # noqa: BLE001
        return {"nome": "", "ocultar": []}
    ocultar = d.get("ocultar") or []
    if not isinstance(ocultar, list):
        ocultar = []
    return {"nome": str(d.get("nome") or ""),
            "ocultar": [str(x).strip().lower() for x in ocultar if str(x).strip()]}


@app.route("/api/ambiente")
def api_ambiente():
    """O frontend usa isto para esconder botões que só existem no app de mesa."""
    e = edicao()
    return jsonify({"web": MODO_WEB, "versao": "2.0",
                    "dev": os.environ.get("CONCRE_DEV") == "1",
                    "edicao": e["nome"], "ocultar": e["ocultar"]})


def _usuario():
    """Identifica o usuário pelo cookie (a web separa as preferências de cada
    um; no app de mesa existe um só e o valor cai em 'default')."""
    u = (request.cookies.get("concre_uid") or "default").strip()
    return "".join(ch for ch in u if ch.isalnum() or ch in "-_")[:40] or "default"


SESSOES_MAX = 6          # quantas sessoes ficam na memoria
SESSOES_TTL = 3600       # segundos sem uso ate' poder descartar


def _limpar_sessoes():
    """Cada recarregar da janela cria uma sessao nova. Sem limpeza, a planilha
    inteira ficava na memoria uma vez por recarregada."""
    agora = time.time()
    for chave in [c for c, v in SESSIONS.items()
                  if agora - v.get("visto", 0) > SESSOES_TTL]:
        SESSIONS.pop(chave, None)
    if len(SESSIONS) >= SESSOES_MAX:
        antigas = sorted(SESSIONS.items(), key=lambda kv: kv[1].get("visto", 0))
        for chave, _ in antigas[:len(SESSIONS) - SESSOES_MAX + 1]:
            SESSIONS.pop(chave, None)


def _session(sid):
    sess = SESSIONS.get(sid)
    if sess is None:
        _limpar_sessoes()
        sess = SESSIONS[sid] = {"sheets": {}, "active": None, "path": None,
                                "demo": []}
    sess["visto"] = time.time()
    return sess


# ──────────────────────────────────────────────────────────────────────────
# Conversão célula → string (replica o formato observado no build original)
#   - datas  → "YYYY-MM-DD"
#   - NaN / None / NaT → ""
#   - float inteiro (8.0) → "8"   |  float real (0.5) → "0.5"
#   - resto → str(v)
# ──────────────────────────────────────────────────────────────────────────
def _cell(v):
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(v, (pd.Timestamp, datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, np.datetime64):
        return pd.Timestamp(v).strftime("%Y-%m-%d")
    if isinstance(v, (float, np.floating)):
        f = float(v)
        return str(int(f)) if f.is_integer() else str(f)
    if isinstance(v, (int, np.integer)):
        return str(int(v))
    return str(v)


def df_to_payload(df):
    """DataFrame → {"headers": [...], "data": [[str, ...], ...]} (igual ao build)."""
    df = df.copy()
    headers = [str(c) for c in df.columns]
    # Datas → ISO antes de extrair (evita numpy.datetime64 solto)
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime("%Y-%m-%d")
    records = df.where(pd.notna(df), None).values.tolist()
    data = [[_cell(v) for v in row] for row in records]
    return {"headers": headers, "data": data}


def payload_to_df(headers, data):
    """{headers, data} → DataFrame (tudo string; usado para exportar/salvar)."""
    return pd.DataFrame(data, columns=headers, dtype=object)


# ──────────────────────────────────────────────────────────────────────────
# Leitura de arquivos (xlsx/xls/csv)
# ──────────────────────────────────────────────────────────────────────────
def read_any(path_or_buffer, filename):
    """Retorna OrderedDict {nome_planilha: DataFrame}. Mantém o dedup de
    cabeçalhos do pandas ("Data 28", "Data 28.1") — igual ao build original."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                df = pd.read_csv(path_or_buffer, sep=None, engine="python",
                                 encoding=enc, dtype=object)
                return {"Plan1": df}
            except Exception:  # noqa: BLE001 -- tenta a próxima codificação
                if hasattr(path_or_buffer, "seek"):
                    path_or_buffer.seek(0)
                continue
        # último recurso
        if hasattr(path_or_buffer, "seek"):
            path_or_buffer.seek(0)
        return {"Plan1": pd.read_csv(path_or_buffer, engine="python", dtype=object)}
    # Excel: todas as planilhas.
    # dtype=object e' obrigatorio: sem ele o pandas "adivinha" e transforma
    # codigos como 007 em 7 e 1E5 em 100000 — corpo de prova e nota fiscal
    # perdiam os zeros a' esquerda.
    sheets = pd.read_excel(path_or_buffer, sheet_name=None, dtype=object)
    return sheets


def load_into_session(sid, sheets, path=None, mode="replace", copia=False):
    """mode='replace' troca a sessão; mode='add' ANEXA as abas do arquivo às
    existentes (permite cruzar planilhas de arquivos diferentes)."""
    sess = _session(sid)
    novos = {name: df_to_payload(df) for name, df in sheets.items()}
    if mode == "add" and sess["sheets"]:
        primeiro = None
        for name, payload in novos.items():
            final, k = name, 2
            while final in sess["sheets"]:
                final = f"{name} ({k})"; k += 1
            sess["sheets"][final] = payload
            if primeiro is None:
                primeiro = final
        sess["active"] = primeiro or sess["active"]
        # 'path' continua apontando para o arquivo base (Salvar não muda de alvo)
    else:
        sess["sheets"] = novos
        sess["active"] = next(iter(sess["sheets"]), None)
        # A marca de "planilha de demonstracao" vale so' para as abas criadas
        # pelo Modo Teste. Ao abrir um arquivo, tudo que entra e' do usuario —
        # sem isto, uma aba dele chamada "Cadastro" herdava a marca e ficava
        # de fora do Salvar, em silencio.
        sess["demo"] = []
        if path:
            sess["path"] = path
            sess["copia"] = bool(copia)
            # quem veio do arquivo: so' estas podem ser apagadas la' no Salvar
            sess["abas_origem"] = list(novos.keys())
    return sess


def sheet_response(sess, extra=None):
    name = sess["active"]
    payload = sess["sheets"].get(name, {"headers": [], "data": []})
    out = {
        "success": True,
        "active_sheet": name,
        "sheet_name": name,
        "sheets": list(sess["sheets"].keys()),
        "data": payload,
    }
    if extra:
        out.update(extra)
    return jsonify(out)


# ──────────────────────────────────────────────────────────────────────────
# Rotas
# ──────────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    # Serve o HTML diretamente (sem Jinja) para não interpretar chaves no markup.
    return send_file(resource_path(os.path.join("templates", "index.html")),
                     mimetype="text/html")


@app.route("/api/upload", methods=["POST"])
def api_upload():
    sid = request.form.get("session_id", "default")
    f = request.files.get("file")
    if not f:
        return jsonify({"success": False, "error": "nenhum arquivo enviado"}), 400
    try:
        # Salva uma cópia em uploads/ (permite "Salvar no arquivo original")
        safe = os.path.basename(f.filename)
        path = os.path.join(UPLOADS_DIR, safe)
        f.save(path)
        sheets = read_any(path, safe)
        mode = request.form.get("mode", "replace")
        # ATENCAO: este 'path' e' a copia dentro de uploads/, NAO o arquivo do
        # usuario. Marcamos como copia para o Salvar nao mentir dizendo que
        # gravou no original (era exatamente o que acontecia ao arrastar).
        sess = load_into_session(sid, sheets, path=path, mode=mode, copia=True)
        return sheet_response(sess, extra={"copia": True})
    except Exception as e:  # noqa: BLE001
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/load_path", methods=["POST"])
def api_load_path():
    if MODO_WEB:
        return _bloqueado_na_web()
    # App nativo: abre o arquivo DIRETO do caminho real (via diálogo do Windows).
    # Assim o "Salvar" grava no próprio arquivo original, não numa cópia.
    body = request.get_json(force=True, silent=True) or {}
    sid = body.get("session_id", "default")
    path = body.get("path") or ""
    if not path or not os.path.isfile(path):
        return jsonify({"success": False, "error": "arquivo não encontrado"}), 400
    try:
        sheets = read_any(path, os.path.basename(path))
        sess = load_into_session(sid, sheets, path=path,
                                 mode=body.get("mode", "replace"))
        return sheet_response(sess, extra={"path": path})
    except Exception as e:  # noqa: BLE001
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/import_merge", methods=["POST"])
def api_import_merge():
    # [reconstruído] anexa as linhas do arquivo ao sheet alvo (mode=append).
    sid = request.form.get("session_id", "default")
    target = request.form.get("target_sheet")
    f = request.files.get("file")
    sess = _session(sid)
    if not f or target not in sess["sheets"]:
        return jsonify({"success": False, "error": "planilha alvo inválida"}), 400
    try:
        sheets = read_any(f.stream, f.filename)
        incoming = next(iter(sheets.values()))
        inc_payload = df_to_payload(incoming)
        cur = sess["sheets"][target]
        # Alinha colunas pelo NOME (ignorando maiúsculas/espaços). Se nenhuma
        # coluna casar, o append geraria só linhas vazias — avisa em vez disso.
        norm = lambda h: str(h).strip().upper()
        idx = {norm(h): i for i, h in enumerate(inc_payload["headers"])}
        casadas = [h for h in cur["headers"] if norm(h) in idx]
        if not casadas:
            return jsonify({
                "success": False,
                "error": ("nenhuma coluna em comum entre as planilhas — "
                          "os cabeçalhos precisam ter os mesmos nomes. "
                          "Para juntar por uma chave, use o botão Cruzar; "
                          "para abrir como outra aba, use Adicionar.")
            }), 400
        # Coluna que so' existe no arquivo importado ENTRA na planilha, em vez
        # de ser jogada fora em silencio (as linhas antigas ficam vazias nela).
        ja_tem = {norm(h) for h in cur["headers"]}
        novas = []
        for h in inc_payload["headers"]:
            if norm(h) not in ja_tem:
                novas.append(h)
                ja_tem.add(norm(h))
        if novas:
            cur["headers"] = list(cur["headers"]) + novas
            for linha in cur["data"]:
                linha.extend([""] * len(novas))

        for row in inc_payload["data"]:
            new_row = []
            for h in cur["headers"]:
                j = idx.get(norm(h))
                new_row.append(row[j] if j is not None and j < len(row) else "")
            cur["data"].append(new_row)
        return jsonify({"success": True, "data": cur,
                        "colunas_casadas": len(casadas),
                        "colunas_novas": novas,
                        "colunas_total": len(cur["headers"]),
                        "linhas_add": len(inc_payload["data"])})
    except Exception as e:  # noqa: BLE001
        return jsonify({"success": False, "error": str(e)}), 500


BACKUP_DIR = os.path.join(app_dir(), "copias")
BACKUP_MAX = 10          # quantas copias por arquivo


def _guardar_copia(path):
    """Guarda o arquivo ANTES de grava-lo por cima. E' a rede de seguranca de
    quem salva errado: da' para voltar a versao anterior sem depender do Excel."""
    try:
        if not path or not os.path.isfile(path):
            return None
        os.makedirs(BACKUP_DIR, exist_ok=True)
        nome = os.path.basename(path)
        base, ext = os.path.splitext(nome)
        base = "".join(ch for ch in base if ch.isalnum() or ch in " -_")[:40].strip() or "planilha"
        carimbo = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S-%f")[:-3]
        destino = os.path.join(BACKUP_DIR, f"{base}__{carimbo}{ext}")
        shutil.copy2(path, destino)

        # mantem so' as ultimas BACKUP_MAX copias DESTE arquivo
        anteriores = sorted(
            (f for f in os.listdir(BACKUP_DIR) if f.startswith(base + "__")),
            reverse=True)
        for velho in anteriores[BACKUP_MAX:]:
            try:
                os.remove(os.path.join(BACKUP_DIR, velho))
            except OSError:
                pass
        return destino
    except Exception:  # noqa: BLE001 -- copia e' proteção extra, nunca impede salvar
        return None


@app.route("/api/copias", methods=["POST"])
def api_copias():
    """Lista as copias guardadas do arquivo aberto (as mais novas primeiro)."""
    if MODO_WEB:
        return _bloqueado_na_web()
    body = request.get_json(force=True, silent=True) or {}
    sess = _session(body.get("session_id", "default"))
    path = sess.get("path") or ""
    if not path:
        return jsonify({"success": True, "copias": []})
    base, _ = os.path.splitext(os.path.basename(path))
    base = "".join(ch for ch in base if ch.isalnum() or ch in " -_")[:40].strip()
    saida = []
    try:
        for f in sorted(os.listdir(BACKUP_DIR), reverse=True):
            if not f.startswith(base + "__"):
                continue
            inteiro = os.path.join(BACKUP_DIR, f)
            marca = f[len(base) + 2:].rsplit(".", 1)[0]
            try:                       # 2026-08-23_18-36-58 -> 23/08/2026 18:36
                d, h = marca.split("_")
                a_, m_, d_ = d.split("-")
                partes = h.split("-")
                hh, mm = partes[0], partes[1]
                ss = partes[2] if len(partes) > 2 else "00"
                # com os segundos da' para diferenciar dois Salvar seguidos
                quando = f"{d_}/{m_}/{a_} às {hh}:{mm}:{ss}"
            except ValueError:
                quando = marca
            saida.append({
                "arquivo": inteiro,
                "quando": quando,
                "tamanho": os.path.getsize(inteiro),
            })
    except OSError:
        pass
    return jsonify({"success": True, "copias": saida, "pasta": BACKUP_DIR})


@app.route("/api/restaurar_copia", methods=["POST"])
def api_restaurar_copia():
    """Volta o arquivo aberto para uma copia anterior (guardando o estado atual
    antes, para o 'voltar atras' tambem ter volta)."""
    if MODO_WEB:
        return _bloqueado_na_web()
    travado = _bloqueado_sem_licenca()
    if travado:
        return travado
    body = request.get_json(force=True, silent=True) or {}
    sid = body.get("session_id", "default")
    sess = _session(sid)
    copia = (body.get("arquivo") or "").strip()
    destino = sess.get("path")
    if not copia or not os.path.isfile(copia):
        return jsonify({"success": False, "error": "copia nao encontrada"}), 400
    if os.path.dirname(os.path.abspath(copia)) != os.path.abspath(BACKUP_DIR):
        return jsonify({"success": False, "error": "caminho invalido"}), 400
    if not destino:
        return jsonify({"success": False, "error": "sem arquivo aberto"}), 400
    try:
        _guardar_copia(destino)          # o estado de agora vira copia tambem
        shutil.copy2(copia, destino)
        sess_novo = load_into_session(sid, read_any(destino, destino), path=destino)
        return sheet_response(sess_novo, extra={"restaurado": True})
    except Exception as e:  # noqa: BLE001
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/join_preview", methods=["POST"])
def api_join_preview():
    """Diz ANTES de confirmar quantas linhas vao achar par. Sem isto, o
    cruzamento era um salto no escuro: so' dava para saber depois de feito."""
    body = request.get_json(force=True, silent=True) or {}
    sess = _session(body.get("session_id", "default"))
    la, lb = body.get("left_sheet"), body.get("right_sheet")
    ka, kb = body.get("left_key"), body.get("right_key")
    if la not in sess["sheets"] or lb not in sess["sheets"]:
        return jsonify({"success": False, "error": "planilha invalida"}), 400
    pa, pb = sess["sheets"][la], sess["sheets"][lb]
    try:
        ia = pa["headers"].index(ka)
        ib = pb["headers"].index(kb)
    except ValueError:
        return jsonify({"success": False, "error": "coluna invalida"}), 400

    def chave(v):
        return str(v).strip().upper()

    direita = {chave(r[ib]) for r in pb["data"] if ib < len(r)}
    total = len(pa["data"])
    casaram = 0
    exemplo = ""
    for r in pa["data"]:
        if ia < len(r) and chave(r[ia]) in direita:
            casaram += 1
            if not exemplo:
                exemplo = str(r[ia]).strip()[:28]
    return jsonify({"success": True, "matched": casaram, "total": total,
                    "exemplo": exemplo})


@app.route("/api/sheets_info", methods=["POST"])
def api_sheets_info():
    # Cabeçalhos de todas as abas da sessão (p/ montar o modal de cruzamento).
    body = request.get_json(force=True, silent=True) or {}
    sess = _session(body.get("session_id", "default"))
    return jsonify({
        "success": True,
        "sheets": {n: p.get("headers", []) for n, p in sess["sheets"].items()},
        "active": sess["active"],
        "path": sess.get("path"),
        # quem e' de demonstracao: o Modo Teste limpa POR ESTA marca, nunca
        # pelo nome — o usuario pode ter uma aba chamada "Cadastro" de verdade
        "demo": list(sess.get("demo") or []),
    })


@app.route("/api/join_sheets", methods=["POST"])
def api_join_sheets():
    # "PROCV automático": traz colunas da planilha B para a A casando pela
    # coluna-chave (merge left). Chave duplicada em B usa a 1ª ocorrência.
    body = request.get_json(force=True, silent=True) or {}
    sess = _session(body.get("session_id", "default"))
    la, lb = body.get("left_sheet"), body.get("right_sheet")
    ka, kb = body.get("left_key"), body.get("right_key")
    cols = body.get("columns") or []
    destino = body.get("destino", "nova")  # 'nova' | 'atual'
    if la not in sess["sheets"] or lb not in sess["sheets"]:
        return jsonify({"success": False, "error": "planilha inválida"}), 400
    pa, pb = sess["sheets"][la], sess["sheets"][lb]
    if ka not in pa["headers"] or kb not in pb["headers"]:
        return jsonify({"success": False, "error": "coluna-chave inválida"}), 400
    cols = [c for c in cols if c in pb["headers"] and c != kb]
    if not cols:
        return jsonify({"success": False, "error": "escolha ao menos uma coluna para trazer"}), 400
    try:
        dfa = payload_to_df(pa["headers"], pa["data"])
        dfb = payload_to_df(pb["headers"], pb["data"])
        # normaliza a chave dos dois lados (trim + sem diferenciar maiúsculas)
        dfa["__k"] = dfa[ka].astype(str).str.strip().str.upper()
        dfb["__k"] = dfb[kb].astype(str).str.strip().str.upper()
        dfb = dfb.drop_duplicates("__k")[["__k"] + cols]
        # evita colisão de nomes: sufixa colunas que já existem em A
        ren = {}
        for c in cols:
            nc, k = c, 2
            while nc in dfa.columns:
                nc = f"{c} ({k})"; k += 1
            if nc != c:
                ren[c] = nc
        if ren:
            dfb = dfb.rename(columns=ren)
        out = dfa.merge(dfb, on="__k", how="left").drop(columns="__k")
        out = out.where(pd.notna(out), "")
        payload = {"headers": [str(c) for c in out.columns],
                   "data": [[str(v) for v in row] for row in out.values.tolist()]}
        matched = int((dfa["__k"].isin(dfb["__k"])).sum()) if len(dfb) else 0
        if destino == "atual":
            sess["sheets"][la] = payload
            sess["active"] = la
        else:
            name, k = f"{la} + {lb}", 2
            while name in sess["sheets"]:
                name = f"{la} + {lb} ({k})"; k += 1
            sess["sheets"][name] = payload
            sess["active"] = name
            # Cruzamento de abas de demonstracao continua sendo demonstracao:
            # o resultado nao pode acabar dentro do arquivo de quem usa.
            demo = set(sess.get("demo") or [])
            if la in demo and lb in demo:
                sess.setdefault("demo", []).append(name)
        return sheet_response(sess, extra={"matched": matched, "total": len(dfa)})
    except Exception as e:  # noqa: BLE001
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/reload", methods=["POST"])
def api_reload():
    if MODO_WEB:
        return _bloqueado_na_web()
    # "Atualizar tudo": relê o arquivo de origem do disco e recarrega a sessão.
    body = request.get_json(force=True, silent=True) or {}
    sid = body.get("session_id", "default")
    sess = _session(sid)
    path = sess.get("path")
    if not path or not os.path.isfile(path):
        return jsonify({"success": False, "error": "sem arquivo de origem para recarregar"}), 400
    try:
        ativo = sess["active"]
        sheets = read_any(path, os.path.basename(path))
        sess = load_into_session(sid, sheets, path=path)
        if ativo in sess["sheets"]:
            sess["active"] = ativo
        return sheet_response(sess, extra={"path": path})
    except Exception as e:  # noqa: BLE001
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/get_sheet", methods=["POST"])
def api_get_sheet():
    body = request.get_json(force=True, silent=True) or {}
    sess = _session(body.get("session_id", "default"))
    name = body.get("sheet_name")
    if name in sess["sheets"]:
        sess["active"] = name
    return sheet_response(sess)


@app.route("/api/new_sheet", methods=["POST"])
def api_new_sheet():
    body = request.get_json(force=True, silent=True) or {}
    sess = _session(body.get("session_id", "default"))
    name = (body.get("sheet_name") or "Nova Planilha").strip() or "Nova Planilha"
    base, n = name, 2
    while name in sess["sheets"]:
        name = f"{base} ({n})"
        n += 1
    sess["sheets"][name] = {"headers": ["A", "B", "C"], "data": [["", "", ""]]}
    sess["active"] = name
    # Aba criada pelo Modo Teste: nunca deve entrar no arquivo do usuario.
    if body.get("demo"):
        sess.setdefault("demo", [])
        if name not in sess["demo"]:
            sess["demo"].append(name)
    return sheet_response(sess)


@app.route("/api/delete_sheet", methods=["POST"])
def api_delete_sheet():
    body = request.get_json(force=True, silent=True) or {}
    sess = _session(body.get("session_id", "default"))
    name = body.get("sheet_name")
    # O Modo Teste pede a limpeza com somente_demo=True. Assim, mesmo que a tela
    # erre o alvo, o servidor recusa apagar uma aba que veio do arquivo.
    if body.get("somente_demo") and name not in (sess.get("demo") or []):
        return jsonify({"success": False,
                        "error": "esta aba nao e' de demonstracao",
                        "sheets": list(sess["sheets"].keys()),
                        "active_sheet": sess["active"]}), 400
    if name in sess["sheets"] and len(sess["sheets"]) > 1:
        sess["sheets"].pop(name, None)
        if name in (sess.get("demo") or []):
            sess["demo"].remove(name)      # nome livre de novo
        if sess["active"] == name:
            sess["active"] = next(iter(sess["sheets"]), None)
    return jsonify({
        "success": True,
        "sheets": list(sess["sheets"].keys()),
        "active_sheet": sess["active"],
    })


@app.route("/api/save_data", methods=["POST"])
def api_save_data():
    # Atualiza o sheet em memória (auto-save do grid).
    body = request.get_json(force=True, silent=True) or {}
    sess = _session(body.get("session_id", "default"))
    name = body.get("sheet_name")
    if not name:
        # sem nome nao ha' o que gravar: devolver "ok" aqui escondia perda de
        # dados (o front achava que salvou).
        return jsonify({"success": False, "error": "sheet_name ausente"}), 400
    sess["sheets"][name] = {
        "headers": body.get("headers", []),
        "data": body.get("data", []),
    }
    sess["active"] = name
    return jsonify({"success": True})


def _excel_sheet_name(name):
    """Nome de aba aceito pelo Excel: sem []:*?/\\ , ≤31 chars, não vazio.
    (openpyxl levanta erro com caracteres inválidos — era fonte de export falho.)"""
    s = "".join(c for c in str(name) if c not in "[]:*?/\\").strip()
    return (s[:31] or "Plan1")


def _safe_sheet_name(name, used):
    """Nome de aba válido p/ Excel e único no arquivo."""
    base = _excel_sheet_name(name)
    out, k = base, 1
    while out in used:
        suf = f"_{k}"
        out = base[: 31 - len(suf)] + suf
        k += 1
    used.add(out)
    return out


@app.route("/api/save_file", methods=["POST"])
def api_save_file():
    if MODO_WEB:
        return _bloqueado_na_web()
    travado = _bloqueado_sem_licenca()
    if travado:
        return travado
    # Grava a sessão de volta no arquivo de origem (ou num caminho novo vindo
    # do "Salvar como" nativo). Excel: escreve TODAS as planilhas.
    body = request.get_json(force=True, silent=True) or {}
    sess = _session(body.get("session_id", "default"))
    new_path = (body.get("path") or "").strip()
    if new_path:
        if not os.path.splitext(new_path)[1]:
            new_path += ".xlsx"
        sess["path"] = new_path
        sess["copia"] = False
    path = sess.get("path")
    if not path or not sess["sheets"]:
        return jsonify({"success": False, "error": "sem arquivo de origem"}), 400
    if sess.get("copia") and not new_path:
        # O arquivo entrou por arrastar/anexar: o app so' tem uma copia e nao
        # sabe onde esta' o original. Antes gravava na copia e dizia "salvo".
        return jsonify({"success": False, "precisa_destino": True,
                        "error": "este arquivo foi aberto por arrastar, entao o app "
                                 "tem so' uma copia dele. Escolha onde gravar."}), 400
    try:
        if path.lower().endswith(".csv"):
            # CSV só comporta uma planilha: grava a ativa.
            name = body.get("sheet_name") or sess["active"]
            if name in (sess.get("demo") or []):
                return jsonify({"success": False,
                                "error": "a planilha aberta e' a de demonstracao "
                                         "do Modo Teste — ela nao vai para o seu "
                                         "arquivo"}), 400
            payload = sess["sheets"].get(name)
            if not payload:
                return jsonify({"success": False, "error": "planilha inválida"}), 400
            copia_anterior = _guardar_copia(path)
            payload_to_df(payload["headers"], payload["data"]).to_csv(
                path, index=False, encoding="utf-8-sig")
        else:
            used = set()
            # As abas de demonstracao do Modo Teste ficam de fora: elas existem
            # so' para o teste e nao podem contaminar a planilha de trabalho.
            demo = set(sess.get("demo") or [])
            reais = {n: p for n, p in sess["sheets"].items() if n not in demo}
            if not reais:
                return jsonify({"success": False,
                                "error": "so' ha planilha de demonstracao aberta"}), 400
            ext = os.path.splitext(path)[1].lower()
            copia_anterior = _guardar_copia(path)   # como estava antes deste Salvar
            if os.path.exists(path) and ext in (".xlsx", ".xlsm"):
                # arquivo ja' existe: gravar POR CIMA, mexendo so' no que mudou
                _gravar_preservando(path, reais, sess.get("abas_origem"))
            else:
                if ext == ".xls":          # openpyxl nao escreve .xls
                    path = path[:-4] + ".xlsx"
                    sess["path"] = path
                with pd.ExcelWriter(path, engine="openpyxl") as w:
                    for sheet_name, payload in reais.items():
                        df = payload_to_df(payload["headers"], payload["data"])
                        # numeros e datas como numeros e datas, nao como texto
                        df = df.map(_tipar) if hasattr(df, "map") else df.applymap(_tipar)
                        df.to_excel(w, sheet_name=_safe_sheet_name(sheet_name, used),
                                    index=False)
        # Arquivo novo (Salvar como): nao havia estado anterior para guardar.
        # Sem isto, o primeiro Salvar deixava a lista de Versoes vazia e parecia
        # que o recurso nao funcionava.
        if not copia_anterior:
            _guardar_copia(path)
        return jsonify({"success": True, "path": path})
    except PermissionError:
        return jsonify({"success": False,
                        "error": "arquivo em uso — feche-o no Excel e salve de novo"}), 500
    except Exception as e:  # noqa: BLE001
        return jsonify({"success": False, "error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────
# Gravacao preservando o arquivo do usuario
# ──────────────────────────────────────────────────────────────────────────
_RE_INT = re.compile(r"^-?\d+$")
_RE_DEC_PONTO = re.compile(r"^-?\d*\.\d+$")
_RE_DEC_VIRG = re.compile(r"^-?\d*,\d+$")
_RE_MILHAR_BR = re.compile(r"^-?\d{1,3}(\.\d{3})+(,\d+)?$")
_RE_DATA_ISO = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")


def _tipar(txt):
    """Texto da grade → valor com o tipo certo para o Excel. Sem isso todo
    numero vira 'numero armazenado como texto' e SOMA/grafico param de somar."""
    if txt is None:
        return None
    s = str(txt).strip()
    if s == "":
        return None
    if _RE_INT.match(s):
        # "007" e' codigo, nao numero: virar 7 perde o zero a' esquerda (o
        # Excel faz igual). "0" sozinho continua sendo numero.
        semSinal = s[1:] if s[:1] == "-" else s
        if len(semSinal) > 1 and semSinal[0] == "0":
            return s
        try:
            return int(s)
        except ValueError:
            return s
    if _RE_DEC_PONTO.match(s):
        return float(s)
    if _RE_DEC_VIRG.match(s):
        return float(s.replace(",", "."))
    if _RE_MILHAR_BR.match(s):
        return float(s.replace(".", "").replace(",", "."))
    m = _RE_DATA_ISO.match(s)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return s
    return s


def _igual_ao_que_o_app_mostra(valor_arquivo, texto_app):
    """O app so' enxerga texto. Se o texto bate com o que ja' esta' na celula,
    a celula NAO e' reescrita — e assim formula, formato e tipo ficam de pe'."""
    a = _cell(valor_arquivo)
    b = "" if texto_app is None else str(texto_app)
    if a == b.strip() or a.strip() == b.strip():
        return True
    na, nb = _tipar(a), _tipar(b)
    if isinstance(na, (int, float)) and isinstance(nb, (int, float)):
        try:
            return abs(float(na) - float(nb)) < 1e-9
        except (TypeError, ValueError):
            return False
    return False


def _gravar_preservando(path, reais, abas_origem):
    """Grava por cima do .xlsx/.xlsm mexendo SO' nas celulas que mudaram.
    Mantem formatacao, largura de coluna, celulas mescladas, paineis
    congelados, validacoes e — o mais importante — as formulas."""
    from openpyxl import load_workbook

    ext = os.path.splitext(path)[1].lower()
    wb = load_workbook(path, keep_vba=(ext == ".xlsm"))

    # O 2o carregamento (valores ja' calculados) custa caro e so' serve para
    # planilha com formula — por isso e' feito na primeira formula encontrada.
    caixa = {"calc": None, "tentou": False}

    def valor_calculado(aba, i, j):
        if not caixa["tentou"]:
            caixa["tentou"] = True
            try:
                caixa["calc"] = load_workbook(path, data_only=True)
            except Exception:  # noqa: BLE001
                caixa["calc"] = None
        c = caixa["calc"]
        if not c or aba not in c.sheetnames:
            return None
        return c[aba].cell(row=i, column=j).value

    usados = set(wb.sheetnames)
    for nome, payload in reais.items():
        headers = payload.get("headers") or []
        data = payload.get("data") or []
        if nome in wb.sheetnames:
            ws = wb[nome]
        else:
            ws = wb.create_sheet(title=_safe_sheet_name(nome, usados))
        linhas = [headers] + list(data)
        for i, linha in enumerate(linhas, start=1):
            for j in range(1, len(headers) + 1):
                novo = linha[j - 1] if j - 1 < len(linha) else ""
                cel = ws.cell(row=i, column=j)
                atual = cel.value
                # Celula com formula: comparar com o VALOR CALCULADO. Se o
                # usuario nao mexeu, a formula continua viva.
                if isinstance(atual, str) and atual.startswith("="):
                    ref = valor_calculado(ws.title, i, j)
                    if _igual_ao_que_o_app_mostra(ref, novo):
                        continue
                elif _igual_ao_que_o_app_mostra(atual, novo):
                    continue
                cel.value = _tipar(novo)

        # sobras de quando a planilha encolheu
        fim = len(data) + 1
        if ws.max_row > fim:
            ws.delete_rows(fim + 1, ws.max_row - fim)
        if ws.max_column > len(headers) and len(headers) > 0:
            ws.delete_cols(len(headers) + 1, ws.max_column - len(headers))

    # abas que vieram do arquivo e o usuario apagou dentro do app
    for nome in list(abas_origem or []):
        if nome not in reais and nome in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb[nome]

    wb.save(path)


def _send_dataframe(df, fmt, base_name):
    """Gera CSV / XLSX / HTML a partir de um DataFrame e devolve como download."""
    if fmt == "csv":
        buf = io.BytesIO(df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig"))
        return send_file(buf, mimetype="text/csv", as_attachment=True,
                         download_name=f"{base_name}.csv")
    if fmt in ("html", "report"):
        html = (
            "<html><head><meta charset='utf-8'><style>"
            "table{border-collapse:collapse;font-family:Arial;font-size:12px}"
            "th,td{border:1px solid #ccc;padding:4px 8px;text-align:left}"
            "th{background:#1d1c18;color:#fff}</style></head><body>"
            f"<h2>{base_name}</h2>"
            + df.to_html(index=False, border=0)
            + "</body></html>"
        )
        buf = io.BytesIO(html.encode("utf-8"))
        return send_file(buf, mimetype="text/html", as_attachment=True,
                         download_name=f"{base_name}.html")
    # xlsx (default)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        # numeros como numeros: senao o Excel abre tudo como texto e nao soma
        saida = df.map(_tipar) if hasattr(df, "map") else df.applymap(_tipar)
        saida.to_excel(w, index=False, sheet_name=_excel_sheet_name(base_name))
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{base_name}.xlsx",
    )


@app.route("/api/export", methods=["POST"])
def api_export():
    travado = _bloqueado_sem_licenca()
    if travado:
        return travado
    body = request.get_json(force=True, silent=True) or {}
    sess = _session(body.get("session_id", "default"))
    name = body.get("sheet_name") or sess["active"]
    fmt = (body.get("format") or "xlsx").lower()
    payload = sess["sheets"].get(name, {"headers": [], "data": []})
    rows = body.get("filtered_data") or payload["data"]
    df = payload_to_df(payload["headers"], rows)
    return _send_dataframe(df, fmt, str(name or "planilha"))


@app.route("/api/export_report_custom", methods=["POST"])
def api_export_report_custom():
    travado = _bloqueado_sem_licenca()
    if travado:
        return travado
    body = request.get_json(force=True, silent=True) or {}
    fmt = (body.get("format") or "xlsx").lower()
    columns = body.get("columns") or []
    data = body.get("data") or []
    title = body.get("title") or "relatorio"
    # Reordena/seleciona as colunas pedidas (data pode trazer linhas completas).
    df_full = payload_to_df(columns, data) if data and len(data[0]) == len(columns) \
        else pd.DataFrame(data)
    if list(df_full.columns) != columns and columns:
        try:
            df_full.columns = columns[: len(df_full.columns)]
        except Exception:  # noqa: BLE001
            pass
    base = f"relatorio_{title}"
    return _send_dataframe(df_full, fmt, base)


@app.route("/api/export_aoa", methods=["POST"])
def api_export_aoa():
    """Gera .xlsx REAL no servidor a partir de uma matriz (lista de listas).

    Existe porque o download via JS (SheetJS) falhava em algumas máquinas —
    aqui o arquivo é escrito com openpyxl, que já vem no executável. Se vier
    'path', grava direto no disco (diálogo nativo); senão devolve o arquivo.
    """
    travado = _bloqueado_sem_licenca()
    if travado:
        return travado
    body = request.get_json(force=True, silent=True) or {}
    aoa = body.get("aoa") or []
    if not isinstance(aoa, list) or not aoa:
        return jsonify({"success": False, "error": "sem dados"}), 400
    nome = _excel_sheet_name(body.get("sheet_name") or "Planilha")
    path = (body.get("path") or "").strip()
    try:
        ncol = max(len(r) for r in aoa)
        linhas = [list(r) + [""] * (ncol - len(r)) for r in aoa]
        df = pd.DataFrame(linhas[1:], columns=[str(c) for c in linhas[0]]) \
            if len(linhas) > 1 else pd.DataFrame(columns=[str(c) for c in linhas[0]])
        # mesma regra do Salvar: numero sai numero, data sai data
        if len(df):
            df = df.map(_tipar) if hasattr(df, "map") else df.applymap(_tipar)
        if path:
            # Na versao web ninguem pode escolher onde gravar no servidor:
            # gravar em caminho livre so' faz sentido no app de mesa.
            if MODO_WEB:
                return _bloqueado_na_web()
            if not os.path.splitext(path)[1]:
                path += ".xlsx"
            with pd.ExcelWriter(path, engine="openpyxl") as w:
                df.to_excel(w, sheet_name=nome, index=False)
            return jsonify({"success": True, "path": path})
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            df.to_excel(w, sheet_name=nome, index=False)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{body.get('base_name') or 'planilha'}.xlsx",
        )
    except PermissionError:
        return jsonify({"success": False,
                        "error": "arquivo em uso — feche-o no Excel e tente de novo"}), 500
    except Exception as e:  # noqa: BLE001
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/salvar_binario", methods=["POST"])
def api_salvar_binario():
    """Grava um arquivo binario (ex.: PNG) no caminho escolhido pelo usuario.
    Existe porque no app nativo o download do navegador nao pergunta onde
    salvar — o arquivo 'sumia'."""
    if MODO_WEB:
        return _bloqueado_na_web()
    travado = _bloqueado_sem_licenca()
    if travado:
        return travado
    import base64
    body = request.get_json(force=True, silent=True) or {}
    path = (body.get("path") or "").strip()
    dados = body.get("base64") or ""
    if not path or not dados:
        return jsonify({"success": False, "error": "caminho ou conteudo ausente"}), 400
    try:
        if "," in dados[:64]:            # remove prefixo data:image/png;base64,
            dados = dados.split(",", 1)[1]
        with open(path, "wb") as fh:
            fh.write(base64.b64decode(dados))
        return jsonify({"success": True, "path": path})
    except Exception as e:  # noqa: BLE001
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/autoteste_arquivo", methods=["POST"])
def api_autoteste_arquivo():
    """Prova, num arquivo descartavel, que o Salvar nao estraga a planilha de
    quem usa: formula, formatacao, data, zero a' esquerda e numero de verdade.
    E' o unico teste que NAO da' para fazer pela tela — so' olhando o .xlsx."""
    if MODO_WEB:
        return _bloqueado_na_web()
    import tempfile
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill

    checks = []

    def anota(nome, ok, detalhe):
        checks.append({"nome": nome, "ok": bool(ok), "detalhe": detalhe})

    pasta = tempfile.mkdtemp(prefix="concre_teste_")
    caminho = os.path.join(pasta, "teste_do_modo_teste.xlsx")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ensaios"
        ws.append(["CP", "DATA", "FCK", "MPA", "SITUACAO"])
        ws.append(["007", datetime.date(2026, 3, 1), 25, 28.4, None])
        ws.append(["008", datetime.date(2026, 3, 2), 25, 22.1, None])
        for r in (2, 3):
            ws.cell(r, 5).value = f'=IF(D{r}>=C{r},"OK","NAO")'
        ws["A1"].font = Font(bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1D1C18")
        ws.column_dimensions["B"].width = 21
        ws.freeze_panes = "A2"
        ws["D2"].number_format = "0.00"
        wb.create_sheet("Nao mexer")["A1"] = "intacto"
        wb.save(caminho)

        sid = "__autoteste_arquivo__"
        SESSIONS.pop(sid, None)
        load_into_session(sid, read_any(caminho, caminho), path=caminho)
        sess = _session(sid)
        grade = sess["sheets"]["Ensaios"]

        anota("Código com zero à esquerda (007) não vira 7",
              grade["data"][0][0] == "007",
              f'a grade mostra "{grade["data"][0][0]}"')

        # o usuario corrige um ensaio e salva
        grade["data"][0][3] = "29,9"
        _guardar_copia(caminho)        # mesmo caminho do botao Salvar
        _gravar_preservando(caminho, sess["sheets"], sess.get("abas_origem"))

        wb2 = load_workbook(caminho)
        w2 = wb2["Ensaios"]
        anota("A correção foi gravada como NÚMERO, não como texto",
              isinstance(w2["D2"].value, (int, float)) and abs(float(w2["D2"].value) - 29.9) < 1e-9,
              f"D2 = {w2['D2'].value!r} ({type(w2['D2'].value).__name__})")
        anota("As fórmulas da planilha continuam vivas",
              isinstance(w2["E2"].value, str) and w2["E2"].value.startswith("="),
              f"E2 = {w2['E2'].value!r}")
        anota("As datas continuam datas",
              isinstance(w2["B2"].value, (datetime.date, datetime.datetime)),
              f"B2 = {w2['B2'].value!r}")
        anota("O que não foi tocado ficou igual",
              w2["D3"].value == 22.1 and w2["A3"].value == "008",
              f"D3 = {w2['D3'].value!r}, A3 = {w2['A3'].value!r}")
        anota("Negrito, cor e largura de coluna continuam iguais",
              bool(w2["A1"].font.bold) and str(w2["A1"].fill.fgColor.rgb).endswith("1D1C18")
              and float(w2.column_dimensions["B"].width) == 21,
              f"largura B = {w2.column_dimensions['B'].width}")
        anota("Painel congelado e formato de número continuam iguais",
              w2.freeze_panes == "A2" and w2["D2"].number_format == "0.00",
              f"congelado em {w2.freeze_panes}, formato {w2['D2'].number_format}")
        anota("As outras abas do arquivo não foram mexidas",
              "Nao mexer" in wb2.sheetnames and wb2["Nao mexer"]["A1"].value == "intacto",
              f"abas: {wb2.sheetnames}")

        # o app tem que ter guardado uma copia ANTES de sobrescrever
        base_copia = "".join(ch for ch in os.path.splitext(os.path.basename(caminho))[0]
                             if ch.isalnum() or ch in " -_")[:40].strip()
        copias = [f for f in os.listdir(BACKUP_DIR)] if os.path.isdir(BACKUP_DIR) else []
        anota("Guardou uma copia do arquivo antes de gravar por cima",
              any(f.startswith(base_copia + "__") for f in copias),
              f"{len([f for f in copias if f.startswith(base_copia + '__')])} copia(s) na pasta 'copias'")

        # a planilha de demonstracao nunca pode entrar no arquivo
        sess["sheets"]["DEMONSTRACAO_FALSA"] = {"headers": ["X"], "data": [["lixo"]]}
        sess.setdefault("demo", []).append("DEMONSTRACAO_FALSA")
        reais = {n: p for n, p in sess["sheets"].items()
                 if n not in set(sess.get("demo") or [])}
        _gravar_preservando(caminho, reais, sess.get("abas_origem"))
        anota("A planilha de demonstração não entra no arquivo",
              "DEMONSTRACAO_FALSA" not in load_workbook(caminho).sheetnames,
              f"abas: {load_workbook(caminho).sheetnames}")

        SESSIONS.pop(sid, None)
    except Exception as e:  # noqa: BLE001
        anota("O teste do arquivo rodou até o fim", False, f"{type(e).__name__}: {e}")
    finally:
        shutil.rmtree(pasta, ignore_errors=True)

    return jsonify({"success": True, "checks": checks,
                    "ok": all(c["ok"] for c in checks)})


# ──────────────────────────────────────────────────────────────────────────
# Mensalidade
# ──────────────────────────────────────────────────────────────────────────
# Onde o app procura versao nova. Ja' vem preenchido para quem recebe o
# programa nao precisar configurar nada; da' para trocar pelas preferencias
# (__url_atualizacao) se um dia o endereco mudar.
# Numa linha so' de proposito: quebrado no meio do caminho, nenhuma busca
# por "usuario/repositorio" encontra este endereco — e trocar de lugar sem
# perceber que um dos arquivos ficou para tras nao da' erro, da' "nao
# consegui verificar", que parece falta de internet.
URL_ATUALIZACAO_PADRAO = "https://raw.githubusercontent.com/embixx/Concrestats/main/atualizacao/manifesto.json"

_LICENCA = {"dados": None, "erro": None, "arquivo": None}


def _prefs_cru():
    try:
        with open(PREFS_FILE, "r", encoding="utf-8") as fh:
            return json.load(fh) or {}
    except Exception:  # noqa: BLE001
        return {}


def _prefs_grava(campos):
    atual = _prefs_cru()
    atual.update(campos)
    try:
        with open(PREFS_FILE, "w", encoding="utf-8") as fh:
            json.dump(atual, fh, ensure_ascii=False, indent=2)
    except OSError:
        pass


def _carregar_licenca():
    """Procura o licenca.key ao lado do executável e confere a assinatura."""
    caminho = licenca.procurar(app_dir())
    if not caminho:
        _LICENCA.update({"dados": None, "erro": None, "arquivo": None})
        return
    dados, erro = licenca.ler_arquivo(caminho)
    _LICENCA.update({"dados": dados, "erro": erro, "arquivo": caminho})


def estado_licenca():
    """Situação atual + marca o relógio, para atrasá-lo não render licença."""
    p = _prefs_cru()
    s = licenca.situacao(_LICENCA["dados"],
                         marca_do_relogio=p.get("__relogio"),
                         inicio_do_teste=p.get("__teste_desde"))
    hoje = datetime.date.today().isoformat()
    novos = {}
    if not p.get("__relogio") or hoje > p["__relogio"]:
        novos["__relogio"] = hoje
    if not _LICENCA["dados"] and not p.get("__teste_desde"):
        novos["__teste_desde"] = hoje
    if novos:
        _prefs_grava(novos)
    if _LICENCA["erro"]:
        s = dict(s)
        s["aviso_arquivo"] = _LICENCA["erro"]
    return s


def _bloqueado_sem_licenca():
    """Vencido não tranca os dados: só fecha gravar e exportar."""
    s = estado_licenca()
    if s.get("pode_gravar"):
        return None
    return jsonify({"success": False, "sem_licenca": True,
                    "estado": s["estado"], "error": s["texto"]}), 402


@app.route("/api/licenca", methods=["GET", "POST"])
def api_licenca():
    if request.method == "GET":
        return jsonify(estado_licenca())
    # POST: instalar um licenca.key que o cliente recebeu
    if MODO_WEB:
        return _bloqueado_na_web()
    body = request.get_json(force=True, silent=True) or {}
    origem = (body.get("path") or "").strip()
    conteudo = body.get("conteudo")
    if conteudo:
        dados, erro = licenca.interpretar(conteudo)
    elif origem:
        dados, erro = licenca.ler_arquivo(origem)
    else:
        return jsonify({"success": False, "error": "informe o arquivo da licença"}), 400
    if erro:
        return jsonify({"success": False, "error": erro}), 400
    try:
        destino = os.path.join(app_dir(), "licenca.key")
        if origem and os.path.abspath(origem) != os.path.abspath(destino):
            shutil.copy2(origem, destino)
        elif conteudo:
            with open(destino, "w", encoding="utf-8") as fh:
                fh.write(conteudo)
        _carregar_licenca()
        return jsonify({"success": True, "situacao": estado_licenca()})
    except OSError as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/pagamento")
def api_pagamento():
    """Dados para a tela de pagamento: planos, codigo PIX e QR.

    Se ninguem configurou a chave PIX, devolve configurado=False e a tela nem
    aparece — melhor nao ter botao de pagar do que ter um que nao recebe.
    """
    try:
        return jsonify(pagamento.cobranca(request.args.get("plano")))
    except Exception as e:  # noqa: BLE001 -- a tela some, o app continua
        return jsonify({"configurado": False, "error": str(e)})


_carregar_licenca()      # lê a licença uma vez, quando o app sobe


def _url_de_atualizacao_segura(url):
    """Devolve None se o endereco pode ser consultado, ou o motivo da recusa."""
    import ipaddress
    import socket
    import urllib.parse
    try:
        u = urllib.parse.urlparse(url)
    except ValueError:
        return "endereco de atualizacao invalido"
    if u.scheme not in ("http", "https"):
        return "so' http ou https"
    if not u.hostname:
        return "endereco sem servidor"
    try:
        infos = socket.getaddrinfo(u.hostname, None)
    except OSError:
        return "servidor nao encontrado"
    for familia, _, _, _, endereco in infos:
        ip = ipaddress.ip_address(endereco[0])
        if (ip.is_private or ip.is_loopback or ip.is_link_local
                or ip.is_reserved or ip.is_multicast):
            return "endereco interno nao e' consultado"
    return None


# ── quem recebe qual versao ─────────────────────────────────────────────────
# Duas maneiras, que servem a coisas diferentes:
#
#   CANAL   e' o do dia a dia. Quem testa fica no canal "teste" e recebe as
#           versoes novas primeiro; o cliente fica no "estavel" e so' recebe o
#           que ja' passou por eles. Cada canal e' um manifesto proprio.
#
#   LIBERADO PARA  e' o controle fino: o manifesto pode trazer uma lista de
#           instalacoes, e so' elas recebem aquela versao. Serve para mandar uma
#           correcao para UMA pessoa antes de soltar para o resto.
#
# Nenhum dos dois e' barreira de seguranca — a assinatura e' que garante que o
# pacote e' nosso. Estes dois decidem so' a QUEM ele e' oferecido. Ainda assim a
# lista viaja dentro do que e' assinado, para ninguem reescrever a mao e
# empurrar uma versao antiga para quem nao devia.
CANAIS = ("estavel", "teste")


def _id_da_instalacao():
    """Identificador curto e estavel desta copia instalada.

    Nao e' o nome de ninguem: sao 10 caracteres sorteados na primeira vez que o
    programa roda, guardados nas preferencias. Serve para o autor dizer 'essa
    versao vai so' para esta instalacao aqui' sem precisar de cadastro, login
    nem servidor.
    """
    prefs = _prefs_cru()
    ident = str(prefs.get("__instalacao") or "").strip()
    if not ident:
        import secrets
        ident = secrets.token_hex(5)
        _prefs_grava({"__instalacao": ident})
    return ident


def _canal_atual():
    c = str(_prefs_cru().get("__canal_atualizacao") or "estavel").strip().lower()
    return c if c in CANAIS else "estavel"


def _url_do_canal(base, canal):
    """No canal de teste o manifesto e' outro arquivo, ao lado do primeiro:
    .../manifesto.json  ->  .../manifesto-teste.json"""
    if canal == "estavel" or not base:
        return base
    if base.endswith(".json"):
        return base[:-5] + "-" + canal + ".json"
    return base


def _mais_nova(anunciada, instalada):
    """A versao anunciada e' realmente mais recente que a instalada?

    Comparar com != aceitava qualquer diferenca, inclusive para tras: bastava
    publicar por engano um manifesto antigo e o programa ofereceria voltar a
    uma versao anterior chamando isso de atualizacao. E comparar como TEXTO
    tambem nao serve — "01/12/2026" vem antes de "29/08/2026" no alfabeto.

    Se as duas datas nao forem legiveis, cai no comportamento antigo (qualquer
    diferenca conta), que e' o mais seguro para versao escrita a mao.
    """
    def data(v):
        try:
            return datetime.datetime.strptime(str(v).strip(), "%d/%m/%Y %H:%M")
        except (ValueError, TypeError):
            return None
    a, i = data(anunciada), data(instalada)
    if a and i:
        return a > i
    return bool(anunciada) and anunciada != instalada


def _liberado_para_mim(info, ident):
    """A versao anunciada vale para esta instalacao? Devolve None se sim, ou o
    motivo de nao. Manifesto sem lista vale para todo mundo (o caso normal)."""
    lista = info.get("liberado_para")
    if not lista:
        return None
    if not isinstance(lista, list):
        return "lista de liberacao invalida no manifesto"
    if ident in [str(x).strip() for x in lista]:
        return None
    return "esta versao foi liberada so' para algumas instalacoes"


@app.route("/api/atualizacao", methods=["GET", "POST"])
def api_atualizacao():
    """Diz se saiu versão nova. A fonte é um endereço configurável em
    prefs.json (__url_atualizacao); sem ele, não incomoda ninguém."""
    # O canal e' da MAQUINA, nao da aba do navegador. Gravar por /api/prefs
    # colocava o valor num balde por usuario, e a leitura acontece na raiz —
    # entao a escolha era salva e ignorada, calada.
    if request.method == "POST":
        if MODO_WEB:
            return _bloqueado_na_web()
        body = request.get_json(force=True, silent=True) or {}
        canal = str(body.get("canal") or "").strip().lower()
        if canal not in CANAIS:
            return jsonify({"success": False,
                            "error": "canal deve ser um de: " + ", ".join(CANAIS)}), 400
        _prefs_grava({"__canal_atualizacao": canal})
        return jsonify({"success": True, "canal": canal})

    atual = "2.0"
    try:
        with open(resource_path(os.path.join("static", "versao.json")),
                  encoding="utf-8") as fh:
            atual = (json.load(fh) or {}).get("versao", atual)
    except Exception:  # noqa: BLE001
        pass
    ident, canal = _id_da_instalacao(), _canal_atual()
    base = _prefs_cru().get("__url_atualizacao") or URL_ATUALIZACAO_PADRAO
    url = _url_do_canal(base, canal)
    comum = {"success": True, "versao_atual": atual, "instalacao": ident,
             "canal": canal, "canais": list(CANAIS)}
    if not url:
        return jsonify(dict(comum, verificou=False))

    # O endereco vem das preferencias, que sao gravaveis pela propria tela.
    # Sem estas travas, alguem podia apontar para file:// (ler arquivo do
    # disco) ou para um endereco interno da rede e usar a resposta como
    # espelho. So' http/https, e nada de endereco privado.
    erro = _url_de_atualizacao_segura(url)
    if erro:
        return jsonify(dict(comum, verificou=False, motivo=erro))
    try:
        import urllib.request
        pedido = urllib.request.Request(url, headers={"User-Agent": "Concrestats"})
        with urllib.request.urlopen(pedido, timeout=6) as r:
            if int(r.headers.get("Content-Length") or 0) > 64 * 1024:
                raise ValueError("resposta grande demais")
            info = json.loads(r.read(64 * 1024).decode("utf-8", "replace"))
        nova = str(info.get("versao", "")).strip()
        # A versao pode existir e mesmo assim nao ser para esta instalacao.
        # Nesse caso a tela nao mente dizendo "voce esta' na mais recente":
        # diz que ha' uma versao restrita, e continua sem oferecer o download.
        restricao = _liberado_para_mim(info, ident)
        return jsonify(dict(comum, verificou=True, versao_nova=nova,
                            tem_nova=bool(nova) and _mais_nova(nova, atual)
                                     and not restricao,
                            restrita=bool(restricao),
                            motivo=restricao or "",
                            novidades=[] if restricao else info.get("novidades", []),
                            onde=info.get("onde", "")))
    except Exception as e:  # noqa: BLE001 -- sem rede não é erro do usuário
        return jsonify(dict(comum, verificou=False, motivo=str(e)[:80]))


@app.route("/api/atualizar", methods=["POST"])
def api_atualizar():
    """Baixa e aplica a atualização anunciada no manifesto.

    A ordem importa: conferir a assinatura ANTES de descompactar. Um pacote de
    origem duvidosa não chega nem a ser aberto."""
    if MODO_WEB:
        return _bloqueado_na_web()
    base = _prefs_cru().get("__url_atualizacao") or URL_ATUALIZACAO_PADRAO
    url = _url_do_canal(base, _canal_atual())
    if not url:
        return jsonify({"success": False, "error": "sem endereço de atualização"}), 400
    erro = _url_de_atualizacao_segura(url)
    if erro:
        return jsonify({"success": False, "error": erro}), 400
    try:
        import urllib.request
        pedido = urllib.request.Request(url, headers={"User-Agent": "Concrestats"})
        with urllib.request.urlopen(pedido, timeout=10) as r:
            manifesto = json.loads(r.read(64 * 1024).decode("utf-8", "replace"))
    except Exception as e:  # noqa: BLE001
        return jsonify({"success": False, "error": f"não consegui ler o manifesto ({e})"}), 502

    # A mesma trava da tela vale aqui. Sem isto, quem chamasse este endereco
    # direto baixaria a versao restrita assim mesmo — a tela seria enfeite.
    restricao = _liberado_para_mim(manifesto, _id_da_instalacao())
    if restricao:
        return jsonify({"success": False, "error": restricao}), 403

    arquivo = manifesto.get("arquivo") or ""
    erro = _url_de_atualizacao_segura(arquivo)
    if erro:
        return jsonify({"success": False, "error": f"endereço do pacote recusado: {erro}"}), 400
    try:
        pacote = atualizador.baixar(arquivo)
    except Exception as e:  # noqa: BLE001
        return jsonify({"success": False, "error": f"falhou ao baixar ({e})"}), 502

    motivo = atualizador.conferir_pacote(pacote, manifesto, licenca.CHAVE_PUBLICA)
    if motivo:
        return jsonify({"success": False, "error": motivo}), 400

    destino = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    ok, msg = atualizador.aplicar(pacote, destino)
    if not ok:
        return jsonify({"success": False, "error": msg}), 500
    return jsonify({"success": True, "mensagem": msg,
                    "versao": manifesto.get("versao", ""),
                    "reiniciar": True})


@app.route("/api/prefs", methods=["GET", "POST"])
def api_prefs():
    # Preferências persistentes (templates/layouts de gráficos, campos fixos do
    # certificado...) num JSON ao lado do executável — sobrevivem a reinstalar,
    # trocar de máquina ou limpar o cache do WebView2.
    u = _usuario()
    if request.method == "GET":
        try:
            with open(PREFS_FILE, "r", encoding="utf-8") as fh:
                todos = json.load(fh) or {}
        except Exception:  # noqa: BLE001
            return jsonify({})
        # Formato antigo (um usuário só) continua funcionando no app de mesa.
        if u == "default" and not isinstance(todos.get(u), dict):
            return jsonify({k: v for k, v in todos.items() if not isinstance(v, dict) or k != "__users__"})
        return jsonify((todos.get("__users__", {}) or {}).get(u, {}))
    body = request.get_json(force=True, silent=True)
    if not isinstance(body, dict):
        return jsonify({"success": False, "error": "esperado objeto JSON"}), 400
    try:
        # Mescla com o que já existe (cada módulo grava só as suas chaves).
        cur = {}
        if os.path.exists(PREFS_FILE):
            try:
                with open(PREFS_FILE, "r", encoding="utf-8") as fh:
                    cur = json.load(fh) or {}
            except Exception:  # noqa: BLE001
                cur = {}
        if u == "default":
            cur.update(body)            # app de mesa: como sempre foi
        else:
            users = cur.setdefault("__users__", {})
            users.setdefault(u, {}).update(body)
        with open(PREFS_FILE, "w", encoding="utf-8") as fh:
            json.dump(cur, fh, ensure_ascii=False, indent=2)
        return jsonify({"success": True})
    except Exception as e:  # noqa: BLE001
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/receitas", methods=["GET", "POST"])
def api_receitas():
    # [reconstruído] persiste as receitas num JSON ao lado do executável.
    if request.method == "GET":
        if os.path.exists(RECEITAS_FILE):
            try:
                with open(RECEITAS_FILE, "r", encoding="utf-8") as fh:
                    return jsonify(json.load(fh))
            except Exception:  # noqa: BLE001
                return jsonify([])
        return jsonify([])
    body = request.get_json(force=True, silent=True)
    try:
        with open(RECEITAS_FILE, "w", encoding="utf-8") as fh:
            json.dump(body if body is not None else [], fh, ensure_ascii=False, indent=2)
        return jsonify({"success": True})
    except Exception as e:  # noqa: BLE001
        return jsonify({"success": False, "error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────
# Inicialização
# ──────────────────────────────────────────────────────────────────────────
def open_browser():
    try:
        webbrowser.open("http://127.0.0.1:5000")
    except Exception:  # noqa: BLE001
        pass


def _run_server():
    app.run(host="127.0.0.1", port=5000, debug=False, use_reloader=False, threaded=True)


if __name__ == "__main__":
    # Executado como app de mesa: libera os recursos locais (abrir por caminho,
    # salvar no arquivo de origem, recarregar). Servido por gunicorn/host, o
    # módulo é só importado e MODO_WEB continua True.
    MODO_WEB = False
    # Flask sobe numa thread em segundo plano.
    threading.Thread(target=_run_server, daemon=True).start()
    # App NATIVO do Windows: abre numa JANELA PRÓPRIA (WebView2), como Spotify —
    # nada de navegador. Se o pywebview não estiver presente, cai pro navegador.
    try:
        import webview  # pywebview
        time.sleep(1.0)  # garante o servidor no ar antes de criar a janela

        class JsApi:
            """Diálogos NATIVOS do Windows (window.pywebview.api.* no frontend).
            Entregam o caminho REAL do arquivo — o Salvar grava no próprio .xlsx."""

            _FT = ("Planilhas (*.xlsx;*.xls;*.csv)", "Todos os arquivos (*.*)")

            def _win(self):
                return webview.windows[0] if webview.windows else None

            def open_file_dialog(self):
                w = self._win()
                if not w:
                    return None
                mode = getattr(webview, "OPEN_DIALOG", None)
                if mode is None:
                    mode = webview.FileDialog.OPEN
                res = w.create_file_dialog(mode, allow_multiple=False,
                                           file_types=self._FT)
                if not res:
                    return None
                return res[0] if isinstance(res, (list, tuple)) else res

            def save_file_dialog(self, default_name="planilha.xlsx"):
                w = self._win()
                if not w:
                    return None
                mode = getattr(webview, "SAVE_DIALOG", None)
                if mode is None:
                    mode = webview.FileDialog.SAVE
                res = w.create_file_dialog(mode, save_filename=str(default_name),
                                           file_types=self._FT)
                if not res:
                    return None
                return res[0] if isinstance(res, (list, tuple)) else res

        webview.create_window(
            "Concrestats", "http://127.0.0.1:5000",
            width=1280, height=820, min_size=(900, 600),
            js_api=JsApi(),
        )

        # private_mode=False + storage_path: o WebView2 usa um perfil PERSISTENTE
        # ao lado do exe (por padrão o pywebview é "anônimo" e o localStorage —
        # templates, layouts, campos fixos — sumia ao fechar o app).
        # Icone da JANELA: sem isto o pywebview extrai o icone de sys.executable
        # — que no modo de desenvolvimento e' o python.exe, e a janela abria com
        # o icone da cobrinha.
        _ico = resource_path("icon.ico")
        webview.start(
            private_mode=False,
            storage_path=os.path.join(app_dir(), "webview_data"),
            icon=_ico if os.path.isfile(_ico) else None,
        )  # bloqueia até fechar a janela
    except Exception:  # noqa: BLE001 -- fallback: navegador
        threading.Timer(1.0, open_browser).start()
        try:
            while True:
                time.sleep(3600)
        except KeyboardInterrupt:
            pass
